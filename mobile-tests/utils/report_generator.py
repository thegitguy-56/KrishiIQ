"""
Builds all required deliverables from the raw pytest-json-report output:
  - Automation_Test_Report.xlsx (Executed/Passed/Failed/Skipped/Metrics/
    Defects/Pass-Rate sheets)
  - execution-report.html
  - dashboard.html
  - summary.md

Run after pytest finishes:
    python utils/report_generator.py --input mobile-tests/reports/execution-results.json --outdir mobile-tests/reports
"""
import argparse
import json
import os
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MODULE_DISPLAY = {
    "test_authentication": "Authentication",
    "test_authorization": "Authorization",
    "test_camera_upload": "Camera / Upload",
    "test_advisory_profile": "Advisory / Profile Management",
    "test_navigation": "Navigation",
    "test_dashboard": "Dashboard",
    "test_forms": "Forms",
    "test_crud": "CRUD Operations",
    "test_input_validation": "Input Validation",
    "test_error_handling": "Error Handling",
    "test_session_management": "Session Management",
    "test_notifications": "Notifications",
    "test_offline_handling": "Offline Handling",
    "test_accessibility": "Accessibility",
    "test_responsive_ui": "Responsive UI",
}

PRIORITY_MARKERS = {"p1": "P1 - Critical", "p2": "P2 - High", "p3": "P3 - Medium", "p4": "P4 - Low"}


def module_from_nodeid(nodeid: str) -> str:
    file_part = nodeid.split("::")[0]
    base = os.path.basename(file_part).replace(".py", "")
    return MODULE_DISPLAY.get(base, base)


def priority_from_keywords(keywords: dict) -> str:
    for marker, label in PRIORITY_MARKERS.items():
        if marker in keywords:
            return label
    return "P3 - Medium"


def load_results(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def build_rows(report: dict):
    rows = []
    for test in report.get("tests", []):
        nodeid = test.get("nodeid", "")
        outcome = test.get("outcome", "unknown")
        keywords = test.get("keywords", {})
        call = test.get("call", {})
        duration = call.get("duration", test.get("setup", {}).get("duration", 0))
        longrepr = ""
        for phase in ("setup", "call", "teardown"):
            phase_data = test.get(phase, {})
            if phase_data.get("outcome") == "failed":
                longrepr = str(phase_data.get("longrepr", ""))[:500]
                break
        rows.append(
            {
                "id": nodeid.split("::")[-1],
                "nodeid": nodeid,
                "module": module_from_nodeid(nodeid),
                "name": nodeid.split("::")[-1].replace("test_", "").replace("_", " ").title(),
                "priority": priority_from_keywords(keywords),
                "status": outcome,
                "duration_s": round(duration, 2) if duration else 0,
                "error": longrepr,
            }
        )
    return rows


def write_excel(rows, outdir: str):
    wb = Workbook()
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, headers):
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    headers = [
        "Test Case ID", "Module", "Test Name", "Priority", "Preconditions",
        "Test Steps", "Test Data", "Expected Result", "Actual Result", "Status",
    ]

    ws_all = wb.active
    ws_all.title = "Executed Tests"
    style_header(ws_all, headers)

    ws_passed = wb.create_sheet("Passed")
    style_header(ws_passed, headers)
    ws_failed = wb.create_sheet("Failed")
    style_header(ws_failed, headers)
    ws_skipped = wb.create_sheet("Skipped")
    style_header(ws_skipped, headers)

    status_map = {"passed": ws_passed, "failed": ws_failed, "skipped": ws_skipped}

    for r in rows:
        record = [
            r["id"],
            r["module"],
            r["name"],
            r["priority"],
            "App installed on emulator; test backend seeded; user session as required by module",
            "See tests/<module_file>.py::" + r["id"] + " for the exact automated step sequence",
            r["id"],
            "Action completes without error / app state matches expected screen or message",
            (r["error"] if r["status"] == "failed" else f"Completed in {r['duration_s']}s"),
            r["status"].upper(),
        ]
        ws_all.append(record)
        target = status_map.get(r["status"])
        if target is not None:
            target.append(record)

    for ws in (ws_all, ws_passed, ws_failed, ws_skipped):
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 24

    # -- Execution Metrics ---------------------------------------------
    ws_metrics = wb.create_sheet("Execution Metrics")
    total = len(rows)
    passed = sum(1 for r in rows if r["status"] == "passed")
    failed = sum(1 for r in rows if r["status"] == "failed")
    skipped = sum(1 for r in rows if r["status"] == "skipped")
    total_duration = sum(r["duration_s"] for r in rows)
    ws_metrics.append(["Metric", "Value"])
    for c in ("A1", "B1"):
        ws_metrics[c].font = header_font
        ws_metrics[c].fill = header_fill
    ws_metrics.append(["Total Executed", total])
    ws_metrics.append(["Passed", passed])
    ws_metrics.append(["Failed", failed])
    ws_metrics.append(["Skipped", skipped])
    ws_metrics.append(["Pass Rate (%)", round((passed / total) * 100, 2) if total else 0])
    ws_metrics.append(["Total Duration (s)", round(total_duration, 2)])
    ws_metrics.append(["Generated At (UTC)", datetime.utcnow().isoformat()])
    ws_metrics.column_dimensions["A"].width = 24
    ws_metrics.column_dimensions["B"].width = 24

    # -- Defect Summary (failures grouped by module) ---------------------
    ws_defects = wb.create_sheet("Defect Summary")
    ws_defects.append(["Module", "Failed Test ID", "Error Summary"])
    for c in ("A1", "B1", "C1"):
        ws_defects[c].font = header_font
        ws_defects[c].fill = header_fill
    for r in rows:
        if r["status"] == "failed":
            ws_defects.append([r["module"], r["id"], (r["error"] or "")[:300]])
    for col in ("A", "B", "C"):
        ws_defects.column_dimensions[col].width = 30

    # -- Pass Rate Summary (per module) ---------------------------------
    ws_rate = wb.create_sheet("Pass Rate Summary")
    ws_rate.append(["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate (%)"])
    for c in ("A1", "B1", "C1", "D1", "E1", "F1"):
        ws_rate[c].font = header_font
        ws_rate[c].fill = header_fill
    by_module = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0, "skipped": 0})
    for r in rows:
        m = by_module[r["module"]]
        m["total"] += 1
        m[r["status"]] = m.get(r["status"], 0) + 1
    for module, stats in sorted(by_module.items()):
        rate = round((stats.get("passed", 0) / stats["total"]) * 100, 2) if stats["total"] else 0
        ws_rate.append([module, stats["total"], stats.get("passed", 0), stats.get("failed", 0), stats.get("skipped", 0), rate])
    for col in ("A", "B", "C", "D", "E", "F"):
        ws_rate.column_dimensions[col].width = 26

    out_path = os.path.join(outdir, "Automation_Test_Report.xlsx")
    wb.save(out_path)
    return out_path, {"total": total, "passed": passed, "failed": failed, "skipped": skipped}


def write_html_reports(rows, stats, outdir: str):
    pass_rate = round((stats["passed"] / stats["total"]) * 100, 2) if stats["total"] else 0

    def row_html(r):
        color = {"passed": "#16A34A", "failed": "#DC2626", "skipped": "#D97706"}.get(r["status"], "#6B7280")
        return (
            f"<tr><td>{r['id']}</td><td>{r['module']}</td><td>{r['name']}</td>"
            f"<td>{r['priority']}</td><td style='color:{color};font-weight:600'>{r['status'].upper()}</td>"
            f"<td>{r['duration_s']}s</td></tr>"
        )

    rows_html = "\n".join(row_html(r) for r in rows)

    exec_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>KrishiIQ Mobile E2E — Execution Report</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;background:#F0FDF4;color:#111}}
h1{{color:#16A34A}}
table{{border-collapse:collapse;width:100%;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.1)}}
th,td{{border:1px solid #E5E7EB;padding:8px 10px;text-align:left;font-size:13px}}
th{{background:#16A34A;color:#fff;position:sticky;top:0}}
.summary{{display:flex;gap:16px;margin-bottom:20px}}
.card{{background:#fff;border-radius:8px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.1)}}
.card b{{font-size:24px;display:block}}
</style></head><body>
<h1>KrishiIQ Mobile E2E — Execution Report</h1>
<p>Generated: {datetime.utcnow().isoformat()} UTC</p>
<div class="summary">
<div class="card">Total<b>{stats['total']}</b></div>
<div class="card">Passed<b style="color:#16A34A">{stats['passed']}</b></div>
<div class="card">Failed<b style="color:#DC2626">{stats['failed']}</b></div>
<div class="card">Skipped<b style="color:#D97706">{stats['skipped']}</b></div>
<div class="card">Pass Rate<b>{pass_rate}%</b></div>
</div>
<table><thead><tr><th>Test Case ID</th><th>Module</th><th>Test Name</th><th>Priority</th><th>Status</th><th>Duration</th></tr></thead>
<tbody>{rows_html}</tbody></table>
</body></html>"""

    with open(os.path.join(outdir, "execution-report.html"), "w", encoding="utf-8") as fh:
        fh.write(exec_html)

    by_module = defaultdict(lambda: {"total": 0, "passed": 0})
    for r in rows:
        by_module[r["module"]]["total"] += 1
        if r["status"] == "passed":
            by_module[r["module"]]["passed"] += 1

    bars = "\n".join(
        f"<div class='bar-row'><span>{m}</span><div class='bar'><div class='fill' "
        f"style='width:{round((s['passed']/s['total'])*100,1) if s['total'] else 0}%'></div></div>"
        f"<span>{s['passed']}/{s['total']}</span></div>"
        for m, s in sorted(by_module.items())
    )

    dashboard_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>KrishiIQ Mobile E2E — Dashboard</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;background:#0F172A;color:#F1F5F9}}
h1{{color:#4ADE80}}
.bar-row{{display:flex;align-items:center;gap:12px;margin:10px 0}}
.bar-row span:first-child{{width:220px}}
.bar-row span:last-child{{width:70px;text-align:right}}
.bar{{flex:1;background:#1E293B;border-radius:6px;overflow:hidden;height:16px}}
.fill{{background:linear-gradient(90deg,#16A34A,#4ADE80);height:100%}}
.big{{font-size:42px;color:#4ADE80;font-weight:700}}
</style></head><body>
<h1>KrishiIQ Mobile E2E — Dashboard</h1>
<p class="big">{pass_rate}% pass rate</p>
<p>{stats['passed']} passed / {stats['failed']} failed / {stats['skipped']} skipped / {stats['total']} total</p>
<h3>By Module</h3>
{bars}
</body></html>"""

    with open(os.path.join(outdir, "dashboard.html"), "w", encoding="utf-8") as fh:
        fh.write(dashboard_html)


def write_summary_md(rows, stats, outdir: str):
    pass_rate = round((stats["passed"] / stats["total"]) * 100, 2) if stats["total"] else 0
    by_module = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0, "skipped": 0})
    for r in rows:
        m = by_module[r["module"]]
        m["total"] += 1
        m[r["status"]] = m.get(r["status"], 0) + 1

    lines = [
        "# KrishiIQ Mobile E2E — Run Summary",
        "",
        f"- **Total executed:** {stats['total']}",
        f"- **Passed:** {stats['passed']}",
        f"- **Failed:** {stats['failed']}",
        f"- **Skipped:** {stats['skipped']}",
        f"- **Pass rate:** {pass_rate}%",
        "",
        "## By Module",
        "",
        "| Module | Total | Passed | Failed | Skipped | Pass Rate |",
        "|---|---|---|---|---|---|",
    ]
    for module, s in sorted(by_module.items()):
        rate = round((s.get("passed", 0) / s["total"]) * 100, 2) if s["total"] else 0
        lines.append(f"| {module} | {s['total']} | {s.get('passed',0)} | {s.get('failed',0)} | {s.get('skipped',0)} | {rate}% |")

    with open(os.path.join(outdir, "summary.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="pytest-json-report output file")
    parser.add_argument("--outdir", required=True)
    args = parser.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    report = load_results(args.input)
    rows = build_rows(report)
    xlsx_path, stats = write_excel(rows, args.outdir)
    write_html_reports(rows, stats, args.outdir)
    write_summary_md(rows, stats, args.outdir)

    print(f"Wrote {xlsx_path}")
    print(f"Total={stats['total']} Passed={stats['passed']} Failed={stats['failed']} Skipped={stats['skipped']}")
    pass_rate = (stats["passed"] / stats["total"]) * 100 if stats["total"] else 0
    print(f"Pass rate: {pass_rate:.2f}%")

    # Exit non-zero if pass rate < 95%, so CI can gate on it explicitly.
    if stats["total"] > 0 and pass_rate < 95.0:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
