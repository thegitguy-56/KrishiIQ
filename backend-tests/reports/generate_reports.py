"""
Generates every report deliverable from the artifacts produced by the
pytest run (tests/artifacts/*.json) and, if present, the k6 load-test
summary (load/k6-summary.json).

Run from the backend-tests/ directory:
    python reports/generate_reports.py

Outputs (written to reports/output/):
    backend-inventory.md
    endpoint-inventory.xlsx
    security-review.md
    executive-summary.md
    performance-report.md
    findings.xlsx
    test-cases.xlsx
    github-step-summary.md   (consumed by the CI workflow)
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from inventory_data import BACKEND_STACK, MODULE_INVENTORY, ENDPOINT_INVENTORY  # noqa: E402

ROOT = Path(__file__).parent.parent
ARTIFACT_DIR = ROOT / "tests" / "artifacts"
LOAD_DIR = ROOT / "load"
OUT_DIR = Path(__file__).parent / "output"
OUT_DIR.mkdir(exist_ok=True)

SEVERITY_ORDER = ["Critical", "High", "Medium", "Low", "Informational"]
SEVERITY_COLORS = {
    "Critical": "C00000",
    "High": "E36C09",
    "Medium": "FFC000",
    "Low": "92D050",
    "Informational": "BFBFBF",
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _load_json(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text())
    except Exception:
        return default


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 80)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Load artifacts
# ---------------------------------------------------------------------------
catalog = _load_json(ARTIFACT_DIR / "test_catalog.json", [])
findings = _load_json(ARTIFACT_DIR / "findings.json", [])
run_summary = _load_json(
    ARTIFACT_DIR / "run_summary.json",
    {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "error": 0, "findings_count": 0, "target": "unknown"},
)
k6_summary = _load_json(LOAD_DIR / "k6-summary.json", None)

generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


# ---------------------------------------------------------------------------
# test-cases.xlsx
# ---------------------------------------------------------------------------
def build_test_cases_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = [
        "Test Case ID", "Category", "Title", "Objective", "Preconditions",
        "Test Steps", "Test Data", "Expected Result", "Severity", "Status",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    by_category = Counter()
    by_status = Counter()

    for row in sorted(catalog, key=lambda r: r["id"]):
        ws.append([
            row["id"], row["category"], row["title"], row["objective"],
            row["preconditions"], row["steps"], row["test_data"],
            row["expected"], row["severity"], row["status"],
        ])
        by_category[row["category"]] += 1
        by_status[row["status"]] += 1
        status_cell = ws.cell(row=ws.max_row, column=10)
        if row["status"] == "Pass":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif row["status"] == "Fail":
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif row["status"] == "Skip":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for cell in ws["A"] + ws["B"] + ws["I"] + ws["J"]:
        cell.alignment = Alignment(vertical="top")
    for col in ("D", "E", "F", "G", "H"):
        for cell in ws[col]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autosize(ws)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    _style_header(ws2, 2)
    ws2.append(["Total test cases", len(catalog)])
    ws2.append(["Passed", by_status.get("Pass", 0)])
    ws2.append(["Failed", by_status.get("Fail", 0)])
    ws2.append(["Skipped", by_status.get("Skip", 0)])
    ws2.append(["Errored", by_status.get("Error", 0)])
    ws2.append(["Target", run_summary.get("target", "unknown")])
    ws2.append(["Generated", generated_at])
    ws2.append([])
    ws2.append(["Category", "Test Case Count"])
    for cat, count in sorted(by_category.items(), key=lambda kv: -kv[1]):
        ws2.append([cat, count])
    _autosize(ws2)

    wb.save(OUT_DIR / "test-cases.xlsx")
    return by_category, by_status


# ---------------------------------------------------------------------------
# findings.xlsx
# ---------------------------------------------------------------------------
def build_findings_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = ["Finding ID", "Severity", "Endpoint", "Description", "Evidence", "Impact", "Remediation", "OWASP", "CWE"]
    ws.append(headers)
    _style_header(ws, len(headers))

    def sev_key(f):
        try:
            return SEVERITY_ORDER.index(f["severity"])
        except ValueError:
            return len(SEVERITY_ORDER)

    for f in sorted(findings, key=sev_key):
        ws.append([
            f["finding_id"], f["severity"], f["endpoint"], f["description"],
            f["evidence"], f["impact"], f["remediation"], f.get("owasp", ""), f.get("cwe", ""),
        ])
        sev_cell = ws.cell(row=ws.max_row, column=2)
        color = SEVERITY_COLORS.get(f["severity"], "FFFFFF")
        sev_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if f["severity"] in ("Critical", "High"):
            sev_cell.font = Font(color="FFFFFF", bold=True)
        for col in ("D", "E", "F", "G"):
            ws.cell(row=ws.max_row, column=ord(col) - 64).alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)

    ws2 = wb.create_sheet("Severity Breakdown")
    ws2.append(["Severity", "Count"])
    _style_header(ws2, 2)
    counts = Counter(f["severity"] for f in findings)
    for sev in SEVERITY_ORDER:
        ws2.append([sev, counts.get(sev, 0)])
    _autosize(ws2)

    wb.save(OUT_DIR / "findings.xlsx")
    return counts


# ---------------------------------------------------------------------------
# endpoint-inventory.xlsx
# ---------------------------------------------------------------------------
def build_endpoint_inventory_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Endpoints"
    headers = ["Method", "Path", "Auth", "Description", "Request Body / Params"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for row in ENDPOINT_INVENTORY:
        ws.append(list(row))
    for col in ("C", "D", "E"):
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)

    ws2 = wb.create_sheet("Modules")
    ws2.append(["Module", "Purpose"])
    _style_header(ws2, 2)
    for row in MODULE_INVENTORY:
        ws2.append(list(row))
    for cell in ws2["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws2)

    wb.save(OUT_DIR / "endpoint-inventory.xlsx")


# ---------------------------------------------------------------------------
# backend-inventory.md
# ---------------------------------------------------------------------------
def build_backend_inventory_md():
    lines = [
        "# KrishiIQ Backend Inventory",
        "",
        f"_Generated {generated_at}_",
        "",
        "## Stack",
        "",
        "| Component | Detail |",
        "|---|---|",
    ]
    for k, v in BACKEND_STACK.items():
        lines.append(f"| {k.replace('_', ' ').title()} | {v} |")

    lines += ["", "## Modules", "", "| Module | Purpose |", "|---|---|"]
    for mod, purpose in MODULE_INVENTORY:
        lines.append(f"| `{mod}` | {purpose} |")

    lines += [
        "",
        "## Endpoints",
        "",
        f"Total routes catalogued: **{len(ENDPOINT_INVENTORY)}** (full detail in `endpoint-inventory.xlsx`).",
        "",
        "| Method | Path | Auth |",
        "|---|---|---|",
    ]
    for method, path, auth, _desc, _params in ENDPOINT_INVENTORY:
        lines.append(f"| {method} | `{path}` | {auth} |")

    (OUT_DIR / "backend-inventory.md").write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# security-review.md
# ---------------------------------------------------------------------------
def build_security_review_md(sev_counts):
    lines = [
        "# KrishiIQ Backend — Security Review",
        "",
        f"_Generated {generated_at} against target `{run_summary.get('target')}`_",
        "",
        "## Methodology",
        "",
        "This review combines:",
        "",
        "1. **SAST** — `bandit` static analysis and `pip-audit` dependency vulnerability scanning, run in CI against the checked-out `backend/` source.",
        "2. **DAST / functional API testing** — a black-box `pytest` + `httpx` suite (this repository) executed against a live, ephemerally-seeded instance of the FastAPI app, covering authentication, authorization, input validation, injection, business logic, configuration, and IDOR/JWT-tampering scenarios.",
        "3. **Load testing** — `k6` baseline (100 VUs / 1 min) and stress (200/500/1000 VUs) runs against the same ephemeral instance.",
        "",
        "All dynamic testing is **non-destructive**: it runs against an ephemeral SQLite database that is seeded and torn down inside the CI job, never against real user data, and never against the production Render deployment unless explicitly opted into (see `README.md`).",
        "",
        "## Severity Summary",
        "",
        "| Severity | Count |",
        "|---|---|",
    ]
    for sev in SEVERITY_ORDER:
        lines.append(f"| {sev} | {sev_counts.get(sev, 0)} |")

    lines += ["", "## Findings", ""]

    def sev_key(f):
        try:
            return SEVERITY_ORDER.index(f["severity"])
        except ValueError:
            return len(SEVERITY_ORDER)

    if not findings:
        lines.append("_No findings were recorded in this run — see `test-cases.xlsx` for full coverage detail._")
    for f in sorted(findings, key=sev_key):
        lines += [
            f"### {f['finding_id']} — {f['severity']} — {f['endpoint']}",
            "",
            f"**Description:** {f['description']}",
            "",
            f"**Evidence:** {f['evidence']}",
            "",
            f"**Impact:** {f['impact']}",
            "",
            f"**Remediation:** {f['remediation']}",
            "",
        ]
        if f.get("owasp") or f.get("cwe"):
            lines.append(f"**Mapping:** {f.get('owasp', '')} {(' / ' + f.get('cwe')) if f.get('cwe') else ''}")
            lines.append("")

    (OUT_DIR / "security-review.md").write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# executive-summary.md
# ---------------------------------------------------------------------------
def build_executive_summary_md(by_category, by_status, sev_counts):
    critical = sev_counts.get("Critical", 0)
    high = sev_counts.get("High", 0)
    total_tests = len(catalog)

    verdict = (
        "**NOT production-ready without remediation** — Critical findings were confirmed."
        if critical > 0
        else ("Ready for production with High-severity items remediated on a fixed timeline." if high > 0
              else "No Critical/High findings confirmed in this run.")
    )

    lines = [
        "# KrishiIQ Backend — Executive Summary",
        "",
        f"_Generated {generated_at}_",
        "",
        f"This report covers an automated security, functional, and performance assessment of the "
        f"KrishiIQ FastAPI backend (`thegitguy-56/KrishiIQ`), executed entirely within GitHub Actions "
        f"against a non-destructive, ephemerally-seeded test instance.",
        "",
        "## Headline numbers",
        "",
        f"- **{total_tests}** structured test cases executed (target: 400+)",
        f"- **{by_status.get('Pass', 0)}** passed, **{by_status.get('Fail', 0)}** failed, **{by_status.get('Skip', 0)}** skipped",
        f"- **{len(findings)}** distinct security/quality findings recorded",
        f"- **{critical}** Critical, **{high}** High severity findings",
        "",
        "## Verdict",
        "",
        verdict,
        "",
        "## Coverage by category",
        "",
        "| Category | Test Cases |",
        "|---|---|",
    ]
    for cat, count in sorted(by_category.items(), key=lambda kv: -kv[1]):
        lines.append(f"| {cat} | {count} |")

    lines += ["", "## Top findings requiring attention", ""]

    def sev_key(f):
        try:
            return SEVERITY_ORDER.index(f["severity"])
        except ValueError:
            return len(SEVERITY_ORDER)

    top = sorted(findings, key=sev_key)[:5]
    if top:
        for f in top:
            lines.append(f"- **[{f['severity']}] {f['finding_id']}** — {f['description']}")
    else:
        lines.append("_None recorded in this run._")

    lines += [
        "",
        "## Where to look next",
        "",
        "- `security-review.md` — full findings detail with evidence and remediation",
        "- `findings.xlsx` — same findings, spreadsheet form for tracking/triage",
        "- `test-cases.xlsx` — every one of the 400+ structured test cases with Pass/Fail status",
        "- `performance-report.md` — k6 load/stress test results",
        "- `backend-inventory.md` / `endpoint-inventory.xlsx` — full API surface catalogue",
    ]

    (OUT_DIR / "executive-summary.md").write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# performance-report.md
# ---------------------------------------------------------------------------
def build_performance_report_md():
    lines = [
        "# KrishiIQ Backend — Performance Report",
        "",
        f"_Generated {generated_at}_",
        "",
    ]

    perf_rows = [r for r in catalog if r["category"] == "Performance"]
    lines += [
        "## In-suite latency smoke tests",
        "",
        f"{len(perf_rows)} functional-level latency/concurrency checks ran as part of the pytest suite "
        f"(individual endpoint budgets, small concurrency bursts, pagination/window-size stress). "
        f"{sum(1 for r in perf_rows if r['status'] == 'Pass')} passed, "
        f"{sum(1 for r in perf_rows if r['status'] == 'Fail')} failed. "
        f"Full detail in `test-cases.xlsx` (Category = Performance).",
        "",
        "## k6 Load / Stress Test",
        "",
    ]

    if not k6_summary:
        lines += [
            "_No `load/k6-summary.json` was found — the k6 job either did not run or was skipped for this "
            "invocation. Re-run the `backend-tests` workflow with the load-test job enabled to populate this "
            "section, or run `k6 run --summary-export=k6-summary.json load/k6-load-test.js` locally against a "
            "non-production target._",
        ]
    else:
        metrics = k6_summary.get("metrics", {})
        http_req_duration = metrics.get("http_req_duration", {}).get("values", {})
        http_reqs = metrics.get("http_reqs", {}).get("values", {})
        http_req_failed = metrics.get("http_req_failed", {}).get("values", {})

        avg = http_req_duration.get("avg")
        mn = http_req_duration.get("min")
        mx = http_req_duration.get("max")
        p95 = http_req_duration.get("p(95)")
        p99 = http_req_duration.get("p(99)")
        rps = http_reqs.get("rate")
        fail_rate = http_req_failed.get("rate")

        lines += [
            "| Metric | Value |",
            "|---|---|",
            f"| Requests/sec | {rps:.2f}" if rps is not None else "| Requests/sec | n/a |",
            f"| Avg response time | {avg:.1f} ms" if avg is not None else "| Avg response time | n/a |",
            f"| Min response time | {mn:.1f} ms" if mn is not None else "| Min response time | n/a |",
            f"| Max response time | {mx:.1f} ms" if mx is not None else "| Max response time | n/a |",
            f"| P95 response time | {p95:.1f} ms" if p95 is not None else "| P95 response time | n/a |",
            f"| P99 response time | {p99:.1f} ms" if p99 is not None else "| P99 response time | n/a |",
            f"| Error rate | {fail_rate * 100:.2f}%" if fail_rate is not None else "| Error rate | n/a |",
            "",
            "### Plain-language interpretation",
            "",
        ]
        if fail_rate is not None and fail_rate > 0.05:
            lines.append(
                f"- Error rate of {fail_rate * 100:.1f}% is **above the 5% threshold** — the backend is "
                f"dropping/rejecting a meaningful share of requests at this load level. Investigate connection "
                f"pool size (`app/database.py` pool_size=10, max_overflow=20), Uvicorn worker count, and whether "
                f"the SQLite/Postgres backing store was the bottleneck."
            )
        else:
            lines.append("- Error rate stayed within an acceptable range for the tested load profile.")
        if p95 is not None:
            if p95 > 2000:
                lines.append(f"- P95 latency of {p95:.0f}ms means 1 in 20 requests took over 2 seconds — likely to feel sluggish to real users under this load.")
            else:
                lines.append(f"- P95 latency of {p95:.0f}ms is within a generally acceptable range for a REST API under load.")

    lines += [
        "",
        "## Load profile executed",
        "",
        "| Stage | Virtual Users | Duration |",
        "|---|---|---|",
        "| Baseline | 100 | 1 minute (sustained) |",
        "| Stress step 1 | 200 | 1 minute |",
        "| Stress step 2 | 500 | 1 minute |",
        "| Stress step 3 | 1000 | 1 minute |",
        "",
        "> Stress stages beyond baseline are **opt-in** in the CI workflow (`RUN_STRESS_TEST=true`) to avoid "
        "accidentally overwhelming a shared or free-tier target. See `README.md` for how to enable them and, "
        "critically, why they should only ever be pointed at the ephemeral CI instance or a target you own and "
        "have explicit permission to load-test.",
    ]

    (OUT_DIR / "performance-report.md").write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# GitHub Actions step summary
# ---------------------------------------------------------------------------
def build_github_step_summary(sev_counts, by_status):
    critical = sev_counts.get("Critical", 0)
    lines = [
        "## 🔎 KrishiIQ Backend Test Report",
        "",
        f"**Target:** `{run_summary.get('target')}`  |  **Generated:** {generated_at}",
        "",
        "### Test cases",
        "",
        "| Total | Passed | Failed | Skipped |",
        "|---|---|---|---|",
        f"| {len(catalog)} | {by_status.get('Pass', 0)} | {by_status.get('Fail', 0)} | {by_status.get('Skip', 0)} |",
        "",
        "### Findings by severity",
        "",
        "| Severity | Count |",
        "|---|---|",
    ]
    for sev in SEVERITY_ORDER:
        lines.append(f"| {sev} | {sev_counts.get(sev, 0)} |")
    lines += [
        "",
        f"{'🔴 **Critical findings present — pipeline configured to fail.**' if critical > 0 else '✅ No Critical findings.'}",
        "",
        "Full reports uploaded as workflow artifacts: `security-review.md`, `executive-summary.md`, "
        "`performance-report.md`, `findings.xlsx`, `test-cases.xlsx`, `backend-inventory.md`, `endpoint-inventory.xlsx`.",
    ]
    (OUT_DIR / "github-step-summary.md").write_text("\n".join(lines))
    return critical


def main():
    by_category, by_status = build_test_cases_xlsx()
    sev_counts = build_findings_xlsx()
    build_endpoint_inventory_xlsx()
    build_backend_inventory_md()
    build_security_review_md(sev_counts)
    build_executive_summary_md(by_category, by_status, sev_counts)
    build_performance_report_md()
    critical = build_github_step_summary(sev_counts, by_status)

    print(f"Reports written to {OUT_DIR}")
    print(f"Total test cases: {len(catalog)} | Findings: {len(findings)} | Critical: {critical}")

    # Exit non-zero only on Critical findings, per the spec's "fail only on Critical" rule.
    if critical > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
