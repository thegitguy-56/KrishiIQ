"""
Report Generator — creates Automation_Test_Report.xlsx, dashboard.html, summary.md
Run after pytest has written execution-results.json.
"""
import json
import os
import sys
from datetime import datetime

# ── openpyxl ──────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed — skipping Excel report")

# ── Add project root to path ──────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
try:
    import config
except ImportError:
    class config:
        REPORTS_DIR     = "reports"
        JSON_REPORT     = "reports/execution-results.json"
        EXCEL_REPORT    = "reports/Automation_Test_Report.xlsx"
        DASHBOARD_HTML  = "reports/dashboard.html"
        SUMMARY_MD      = "reports/summary.md"
        BASE_URL        = ""

# ── Load results ──────────────────────────────────────────────────────────────

def load_results() -> dict:
    path = config.JSON_REPORT
    if not os.path.exists(path):
        print(f"No results file found at {path}")
        return {"total": 0, "passed": 0, "failed": 0, "skipped": 0,
                "run_at": datetime.now().isoformat(), "tests": []}
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


# ── Excel Report ──────────────────────────────────────────────────────────────

GREEN  = "FF1E8449"
RED    = "FFC0392B"
ORANGE = "FFD35400"
BLUE   = "FF1A5276"
LIGHT_GREEN = "FFD5F5E3"
LIGHT_RED   = "FFFDE8E8"
LIGHT_ORANGE= "FFFEF9E7"
GRAY   = "FFF2F3F4"
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
BODY_FONT   = Font(size=10)

def _hdr(ws, cells_values: dict, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    for cell_ref, val in cells_values.items():
        c = ws[cell_ref]
        c.value = val
        c.font  = HEADER_FONT
        c.fill  = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _row_fill(ws, row: int, cols: int, hex_color: str):
    fill = PatternFill("solid", fgColor=hex_color)
    for col in range(1, cols + 1):
        ws.cell(row, col).fill = fill

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


MODULE_MAP = {
    "test_authentication":              "Authentication",
    "test_authorization":               "Authorization",
    "test_navigation":                  "Navigation",
    "test_ui_validation":               "UI Validation",
    "test_forms_and_input":             "Forms / Input Validation",
    "test_crud_operations":             "CRUD Operations",
    "test_error_session_a11y_responsive": "Error / Session / A11y / Responsive",
}


def create_excel(data: dict):
    if not HAS_OPENPYXL:
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tests   = data.get("tests", [])
    passed  = [t for t in tests if t["status"] == "PASSED"]
    failed  = [t for t in tests if t["status"] == "FAILED"]
    skipped = [t for t in tests if t["status"] == "SKIPPED"]

    # ── Sheet: Executed Tests ──────────────────────────────────────────────────
    ws = wb.create_sheet("Executed Tests")
    headers = ["#", "Test ID", "Module", "Markers", "Status", "Duration (s)"]
    _hdr(ws, {get_column_letter(i+1)+"1": h for i, h in enumerate(headers)}, BLUE)
    for i, t in enumerate(tests, 2):
        module = MODULE_MAP.get(t.get("module","").replace("tests.",""), t.get("module",""))
        ws.cell(i,1).value = i-1
        ws.cell(i,2).value = t["id"].split("::")[-1]
        ws.cell(i,3).value = module
        ws.cell(i,4).value = ", ".join(t.get("markers",[]))
        ws.cell(i,5).value = t["status"]
        ws.cell(i,6).value = t.get("duration", 0)
        color = LIGHT_GREEN if t["status"]=="PASSED" else (LIGHT_RED if t["status"]=="FAILED" else LIGHT_ORANGE)
        _row_fill(ws, i, 6, color)
    _auto_width(ws)

    # ── Sheet: Passed ──────────────────────────────────────────────────────────
    for sheet_name, test_list, color in [
        ("Passed",  passed,  LIGHT_GREEN),
        ("Failed",  failed,  LIGHT_RED),
        ("Skipped", skipped, LIGHT_ORANGE),
    ]:
        ws2 = wb.create_sheet(sheet_name)
        _hdr(ws2, {"A1":"#","B1":"Test ID","C1":"Module","D1":"Duration (s)"}, GREEN if sheet_name=="Passed" else (RED if sheet_name=="Failed" else ORANGE))
        for i, t in enumerate(test_list, 2):
            ws2.cell(i,1).value = i-1
            ws2.cell(i,2).value = t["id"].split("::")[-1]
            ws2.cell(i,3).value = MODULE_MAP.get(t.get("module",""), t.get("module",""))
            ws2.cell(i,4).value = t.get("duration", 0)
            _row_fill(ws2, i, 4, color)
        _auto_width(ws2)

    # ── Sheet: Execution Metrics ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Execution Metrics")
    _hdr(ws3, {"A1":"Metric","B1":"Value"}, BLUE)
    total   = data.get("total", 0)
    p_count = data.get("passed", 0)
    f_count = data.get("failed", 0)
    s_count = data.get("skipped", 0)
    pass_rate = round(p_count/total*100, 2) if total else 0
    metrics = [
        ("Run At",        data.get("run_at","")),
        ("Base URL",      data.get("base_url","")),
        ("Total Tests",   total),
        ("Passed",        p_count),
        ("Failed",        f_count),
        ("Skipped",       s_count),
        ("Pass Rate (%)", pass_rate),
        ("Total Duration (s)", round(sum(t.get("duration",0) for t in tests),2)),
    ]
    for i, (k, v) in enumerate(metrics, 2):
        ws3.cell(i,1).value = k; ws3.cell(i,1).font = Font(bold=True)
        ws3.cell(i,2).value = v
    _auto_width(ws3)

    # ── Sheet: Defect Summary ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Defect Summary")
    _hdr(ws4, {"A1":"#","B1":"Defect / Test ID","C1":"Module","D1":"Severity"}, RED)
    for i, t in enumerate(failed, 2):
        ws4.cell(i,1).value = i-1
        ws4.cell(i,2).value = t["id"].split("::")[-1]
        ws4.cell(i,3).value = MODULE_MAP.get(t.get("module",""), t.get("module",""))
        markers = t.get("markers",[])
        severity = "HIGH" if "high" in markers else ("MEDIUM" if "medium" in markers else "LOW")
        ws4.cell(i,4).value = severity
        _row_fill(ws4, i, 4, LIGHT_RED)
    _auto_width(ws4)

    os.makedirs(os.path.dirname(config.EXCEL_REPORT), exist_ok=True)
    wb.save(config.EXCEL_REPORT)
    print(f"✅ Excel report saved: {config.EXCEL_REPORT}")


# ── Dashboard HTML ────────────────────────────────────────────────────────────

DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>KrishiIQ Test Dashboard</title>
<style>
  :root {{--pass:#1e8449;--fail:#c0392b;--skip:#d35400;--bg:#0f172a;--card:#1e293b;--text:#e2e8f0;--muted:#94a3b8}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}}
  header{{background:linear-gradient(135deg,#1e8449,#27ae60);padding:2rem;text-align:center}}
  header h1{{font-size:2rem;font-weight:800;letter-spacing:-0.5px}}
  header p{{color:rgba(255,255,255,.8);margin-top:.5rem}}
  .container{{max-width:1200px;margin:0 auto;padding:2rem}}
  .grid4{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:1rem;margin:2rem 0}}
  .kpi{{background:var(--card);border-radius:1rem;padding:1.5rem;text-align:center;border:1px solid #334155}}
  .kpi .value{{font-size:3rem;font-weight:800;line-height:1}}
  .kpi .label{{font-size:.875rem;color:var(--muted);margin-top:.5rem;text-transform:uppercase;letter-spacing:1px}}
  .kpi.pass .value{{color:var(--pass)}} .kpi.fail .value{{color:var(--fail)}}
  .kpi.skip .value{{color:var(--skip)}} .kpi.total .value{{color:#60a5fa}}
  .progress-bar{{background:#1e293b;border-radius:999px;height:2rem;overflow:hidden;margin:1rem 0;border:1px solid #334155}}
  .progress-fill{{height:100%;background:linear-gradient(90deg,#1e8449,#27ae60);display:flex;align-items:center;justify-content:flex-end;padding:0 1rem;color:#fff;font-weight:700;border-radius:999px;transition:width 1s ease}}
  table{{width:100%;border-collapse:collapse;margin-top:1rem}}
  th{{background:#1e293b;padding:1rem;text-align:left;border-bottom:2px solid #334155;color:var(--muted);font-size:.8rem;text-transform:uppercase;letter-spacing:1px}}
  td{{padding:.75rem 1rem;border-bottom:1px solid #1e293b;font-size:.875rem}}
  tr:hover td{{background:#1e293b}}
  .badge{{padding:.2rem .6rem;border-radius:999px;font-size:.75rem;font-weight:700}}
  .badge-pass{{background:#052e16;color:#86efac}} .badge-fail{{background:#450a0a;color:#fca5a5}}
  .badge-skip{{background:#431407;color:#fdba74}}
  .section{{background:var(--card);border-radius:1rem;padding:1.5rem;margin-bottom:1.5rem;border:1px solid #334155}}
  .section h2{{font-size:1.1rem;font-weight:700;margin-bottom:1rem;color:#60a5fa}}
  .meta{{color:var(--muted);font-size:.85rem;margin-bottom:.25rem}}
  footer{{text-align:center;padding:2rem;color:var(--muted);font-size:.8rem}}
</style>
</head>
<body>
<header>
  <h1>🌾 KrishiIQ — Selenium Test Dashboard</h1>
  <p>Automated E2E Test Execution Report</p>
</header>
<div class="container">
  <div class="section">
    <div class="meta">🕐 Run At: {run_at}</div>
    <div class="meta">🔗 Base URL: {base_url}</div>
  </div>

  <div class="grid4">
    <div class="kpi total"><div class="value">{total}</div><div class="label">Total Tests</div></div>
    <div class="kpi pass"><div class="value">{passed}</div><div class="label">Passed</div></div>
    <div class="kpi fail"><div class="value">{failed}</div><div class="label">Failed</div></div>
    <div class="kpi skip"><div class="value">{skipped}</div><div class="label">Skipped</div></div>
  </div>

  <div class="section">
    <h2>📊 Pass Rate: {pass_rate}%</h2>
    <div class="progress-bar">
      <div class="progress-fill" style="width:{pass_rate}%">{pass_rate}%</div>
    </div>
  </div>

  <div class="section">
    <h2>❌ Failed Tests</h2>
    {failed_table}
  </div>

  <div class="section">
    <h2>✅ All Tests</h2>
    {all_tests_table}
  </div>
</div>
<footer>Generated by KrishiIQ Selenium Framework · {run_at}</footer>
</body>
</html>"""


def _badge(status):
    cls = {"PASSED":"pass","FAILED":"fail","SKIPPED":"skip"}.get(status,"")
    return f'<span class="badge badge-{cls}">{status}</span>'

def _table(tests, cols=None):
    if cols is None:
        cols = ["#", "Test ID", "Module", "Status", "Duration"]
    rows_html = ""
    for i, t in enumerate(tests, 1):
        module = MODULE_MAP.get(t.get("module",""), t.get("module",""))
        rows_html += (
            f"<tr><td>{i}</td>"
            f"<td>{t['id'].split('::')[-1]}</td>"
            f"<td>{module}</td>"
            f"<td>{_badge(t['status'])}</td>"
            f"<td>{t.get('duration',0):.2f}s</td></tr>"
        )
    header = "".join(f"<th>{c}</th>" for c in cols)
    return f"<table><thead><tr>{header}</tr></thead><tbody>{rows_html}</tbody></table>"


def create_dashboard(data: dict):
    tests   = data.get("tests", [])
    failed  = [t for t in tests if t["status"] == "FAILED"]
    total   = data.get("total", 0)
    passed  = data.get("passed", 0)
    f_count = data.get("failed", 0)
    s_count = data.get("skipped", 0)
    pass_rate = round(passed/total*100, 1) if total else 0

    html = DASHBOARD_TEMPLATE.format(
        run_at=data.get("run_at",""),
        base_url=data.get("base_url",""),
        total=total, passed=passed, failed=f_count, skipped=s_count,
        pass_rate=pass_rate,
        failed_table=_table(failed) if failed else "<p style='color:#94a3b8'>No failures 🎉</p>",
        all_tests_table=_table(tests[:200]),
    )
    os.makedirs(os.path.dirname(config.DASHBOARD_HTML), exist_ok=True)
    with open(config.DASHBOARD_HTML, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"✅ Dashboard saved: {config.DASHBOARD_HTML}")


# ── Summary Markdown ──────────────────────────────────────────────────────────

def create_summary(data: dict):
    total     = data.get("total", 0)
    passed    = data.get("passed", 0)
    failed    = data.get("failed", 0)
    skipped   = data.get("skipped", 0)
    pass_rate = round(passed/total*100, 2) if total else 0
    run_at    = data.get("run_at","")

    failed_tests = [t for t in data.get("tests",[]) if t["status"]=="FAILED"]

    md = f"""# KrishiIQ Selenium Test Execution Summary

| Metric | Value |
|--------|-------|
| Run At | {run_at} |
| Base URL | {data.get('base_url','')} |
| Total Tests | {total} |
| ✅ Passed | {passed} |
| ❌ Failed | {failed} |
| ⏭️ Skipped | {skipped} |
| Pass Rate | **{pass_rate}%** |

## Pass Rate: {pass_rate}%

{'✅ Pass rate is above 95% threshold.' if pass_rate >= 95 else '❌ Pass rate is BELOW 95% threshold!'}

## Module Breakdown

| Module | Tests |
|--------|-------|
"""
    from collections import Counter
    module_counts = Counter(
        MODULE_MAP.get(t.get("module",""),t.get("module",""))
        for t in data.get("tests",[])
    )
    for mod, count in sorted(module_counts.items()):
        md += f"| {mod} | {count} |\n"

    if failed_tests:
        md += "\n## Failed Tests\n\n"
        for t in failed_tests[:50]:
            md += f"- `{t['id'].split('::')[-1]}`\n"

    os.makedirs(os.path.dirname(config.SUMMARY_MD), exist_ok=True)
    with open(config.SUMMARY_MD, "w", encoding="utf-8") as fh:
        fh.write(md)
    print(f"✅ Summary saved: {config.SUMMARY_MD}")
    return pass_rate


# ── GitHub Actions Summary ────────────────────────────────────────────────────

def create_github_summary(data: dict):
    total     = data.get("total", 0)
    passed    = data.get("passed", 0)
    failed    = data.get("failed", 0)
    skipped   = data.get("skipped", 0)
    pass_rate = round(passed/total*100, 2) if total else 0

    gh_summary_file = os.environ.get("GITHUB_STEP_SUMMARY", "")
    if not gh_summary_file:
        return

    content = f"""## 🌾 KrishiIQ Selenium Test Results

| | Count |
|---|---|
| ✅ Passed | **{passed}** |
| ❌ Failed | **{failed}** |
| ⏭️ Skipped | **{skipped}** |
| 📊 Total | **{total}** |
| 🎯 Pass Rate | **{pass_rate}%** |

{'> ✅ **PASS** — Pass rate ({pass_rate}%) is above 95% threshold.' if pass_rate >= 95 else f'> ❌ **FAIL** — Pass rate ({pass_rate}%) is BELOW 95% threshold!'}
"""
    with open(gh_summary_file, "a", encoding="utf-8") as fh:
        fh.write(content)
    print("✅ GitHub Actions summary written")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    data = load_results()
    create_excel(data)
    create_dashboard(data)
    pass_rate = create_summary(data)
    create_github_summary(data)
    print(f"\n{'='*50}")
    print(f"TOTAL: {data['total']} | PASSED: {data['passed']} | "
          f"FAILED: {data['failed']} | SKIPPED: {data['skipped']} | "
          f"PASS RATE: {pass_rate}%")
    if pass_rate < 95:
        print("❌ Pass rate below 95% threshold!")
        sys.exit(1)
    else:
        print("✅ Pass rate above 95% threshold!")
        sys.exit(0)
